import re
from tkinter import *
from openpyxl import load_workbook

default_list = ['Передача права на использование ПО ViPNet Administrator 4.х (КС2)',
                'Сертификат активации сервиса совместной технической поддержки ПО ViPNet Policy Manager 4.x на срок 1'
                'год, уровень - Стандартный',
                'ПАК ViPNet Coordinator HW100 C 4.x (WiFi) 12,00'
                ]

# OPENPYXL CODE
my_table = load_workbook('infotecs.xlsx')
sheet_name = (str(my_table.active))[12:-2]
sheet = my_table[sheet_name]
rows = sheet.max_row
cols = sheet.max_column

for i in range(0, len(default_list)):
    result = re.search(",00", default_list[i])
    print(result)


def get_list(value):
    my_list = []
    for j in range(1, rows + 1):
        if sheet.cell(row=j, column=value).value is None:
            j = j + 1
        else:
            my_list.append(sheet.cell(row=j, column=value).value)
    print(my_list)
    return my_list


def_list = get_list(6)
print(def_list)

def save_list():
    f = open('result.xpsx', 'w')
    f.writelines("\n".join(box.get(0, END)))
    f.close()

def calculate():
    a = Toplevel()
    a.title("Рассчет стоимости")
    name_label = Label(a, text="Введите имя:").pack()
    entry = Entry(a, width=50)
    entry.pack()
    surname_label = Label(a, text="Введите фамилию:").pack()
    entry2 = Entry(a, width=50)
    entry2.pack()
    button1 = Button(a, text="Найти", command=save_list) \
        .pack(fill=X)


root = Tk()
root.title("Главное окно")
Button(text="Button", width=20).pack()
Label(text="Label", width=20, height=3) \
    .pack()
Button(text="About", width=20, command=calculate) \
    .pack()

root.mainloop()
