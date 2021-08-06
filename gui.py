from tkinter import *
import re
import openpyxl
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import tkinter.messagebox as mb
import xlsxwriter

# OPENPYXL CODE
my_table = load_workbook('infotecs_test.xlsx')
sheet_name = (str(my_table.active))[12:-2]
sheet = my_table[sheet_name]
rows = sheet.max_row
cols = sheet.max_column
wb = openpyxl.Workbook()
data = []


def save_excel():
    print(data)
    for j in range(len(data)):
        data[j].reverse()
        x = j + 1
        data[j].append(x)
        data[j].reverse()
    my_table.close()
    workbook = xlsxwriter.Workbook('result.xlsx')
    worksheet = workbook.add_worksheet()
    cell_format = workbook.add_format({'bold': True})
    cell_format.set_font_size(20)
    text_format = workbook.add_format({'text_wrap': True})
    worksheet.write(0, 0, 'Рассчет стоимости для заказчика', cell_format)
    print(data, 'save excel')
    len_data = len(data)
    worksheet.add_table(1, 0, len_data + 1, 4, {
        'data': data,
        'columns': [{'header': '№ п/п'},
                    {'header': 'Перечень СКЗИ'},
                    {'header': 'Количество, ед.'},
                    {'header': 'Цена за ед., руб'},
                    {'header': 'Сумма, руб.'},
                    ]})
    some_str = '=SUM(E3:E' + str(len_data + 2) + ')'
    worksheet.write_formula((len_data + 3), 4, some_str)
    worksheet.write((len_data + 3), 0, 'Итоговая сумма: ', cell_format)
    workbook.close()


def get_list_of_ids():
    my_list = []
    tmp = []
    for i in range(1, rows + 1):
        if sheet.cell(row=i, column=2).value is None:
            i = i + 1
        else:
            my_list.append(sheet.cell(row=i, column=2).value)
    pattern_valid = '\D{1,3}-\d{1,3}'
    for i in my_list:
        result = re.search(pattern_valid, i)
        if result is not None:
            tmp.append(i)
    for i in tmp:
        print(i)
    return result


def get_parent(some_list: list):
    child_list = []
    for i in range(len(some_list)):
        for j in range(1, rows + 1):
            if some_list[i] == sheet.cell(row=j, column=2).value:
                pattern = re.compile('\d{3}')
                result = pattern.search(str(sheet.cell(row=j, column=5).value))
                if result is not None:
                    child_list.append(some_list[i])
    return child_list


# creates a list with sales indices
def get_list_of_sales():
    my_list = []
    for i in range(1, rows + 1):
        if sheet.cell(row=i, column=5).value is None:
            i = i + 1
        else:
            my_list.append(sheet.cell(row=i, column=5).value)

    return my_list


# removes some excel trash
def parse_list_of_ids(list_sample: list):
    len_list = len(list_sample)
    for i in reversed(range(len_list)):
        result = re.search('\d{2}', list_sample[i])
        if result is None:
            list_sample.pop(i)
    len_list = len(list_sample)
    for i in reversed(range(len_list)):
        result1 = re.search(re.escape('*'), list_sample[i])
        if result1 is not None:
            list_sample.pop(i)
    len_list = len(list_sample)
    for i in reversed(range(len_list)):
        result1 = re.search(re.escape('960,00'), list_sample[i])
        if result1 is not None:
            list_sample.pop(i)
    list_sample.pop(0)
    return list_sample


# finds children for parent by a single product id. returns dict
def find_childs_by_id(product_id):
    list_sample = {}
    for i in range(1, rows + 1):
        pattern = re.compile(re.escape(product_id), flags=re.IGNORECASE)
        result = pattern.findall(str(sheet.cell(row=i, column=2).value))
        try:
            if result[0] is not None:
                list_sample.update({str(sheet.cell(row=i, column=2).value): product_id})
        except:
            pass
    return list_sample


# creating a dict with child-parents links
def get_dict_of_parents(list_of_id: list):
    list_of_id = get_parent(list_of_id)
    full_dict = {}
    for i in range(len(list_of_id)):
        update = find_childs_by_id(list_of_id[i])
        full_dict.update(update)
    return full_dict


# finds a parent in child-parent dict
def find_parent(child_name: str, dict_sampl: dict):
    parent = dict_sampl.get(child_name)
    return parent


# turns on another window
def calculate():
    a.deiconify()


# calculate taxes modifier (НДС)
def calculate_taxes(name):
    modifier = 0
    for j in range(1, rows + 1):
        if sheet.cell(row=j, column=3).value == name:
            if sheet.cell(row=j, column=6).value == 'нет':
                modifier = 1
            else:
                modifier = 1
    return modifier


# function finds item in top table
def find_item():
    box2.delete(0, END)
    for j in range(box.size()):
        try:
            pattern = re.compile(re.escape(entry.get()), flags=re.IGNORECASE)
            result = pattern.findall(box.get(j))
            if result[0] is not None:
                box2.insert(END, box.get(j))
        except:
            pass
    entry.delete(0, END)


# save data to data[]
def save_data():
    price = 0
    final_price = 0
    price_modifier = 1
    taxes_modifier = 1
    amount = int(entry_calculate.get())
    name = str((box2.get(ACTIVE)))
    # getting price modifier
    for j in range(1, rows + 1):
        if sheet.cell(row=j, column=3).value == name:
            print('вхождение в вычисление модификатора цены')
            # parent_id = find_parent(str(sheet.cell(row=j, column=2).value))
            if re.search('\d{3}', str(sheet.cell(row=j, column=5).value)) is None:
                if re.search('%', str(sheet.cell(row=j, column=5).value)) is not None:
                    price_modifier = str(sheet.cell(row=j, column=5).value)[0] \
                                     + str(sheet.cell(row=j, column=5).value)[1]
                    price_modifier = int(price_modifier) / 100
                    print('преобразовали модификатор из хх% в 0.х')
                else:
                    print(sheet.cell(row=j, column=5).value)
                    price_modifier = float(sheet.cell(row=j, column=5).value)
                    print('взяли модификатор в формате 0.х')
            else:
                print('нету модификатора цены, это родительский продукт')
                price_modifier = 1
    print('итоговый модификатор цены после циклов вычислений', price_modifier)
    # getting parent price or price with amount modifier
    for j in range(1, rows + 1):
        if sheet.cell(row=j, column=3).value == name:
            parent_id = find_parent(str(sheet.cell(row=j, column=2).value), dict_sampl=dict_sample)
            if sheet.cell(row=j, column=4).value == 'не предусмотрена':
                for i in range(1, rows + 1):
                    if sheet.cell(row=i, column=2).value == parent_id:
                        print(parent_id, 'parent ID')
                        price = int(sheet.cell(row=i, column=5).value)
                taxes_modifier = calculate_taxes(name)
                print(taxes_modifier, '- налоговый модификатор')
                print(amount, '- количество')
                print(price_modifier, '- модификатор цены')
                print(price, '- цена')
                final_price = amount * price_modifier * price
            if sheet.cell(row=j, column=4).value != 'не предусмотрена':
                for i in range(1, rows + 1):
                    if sheet.cell(row=i, column=2).value == parent_id:
                        print(sheet.cell(row=i, column=3).value)
                        print(parent_id, 'parent_id')
                        z = i
                        while re.search(re.escape('>'), sheet.cell(row=z, column=4).value) is None:
                            result = re.findall('[+]?\d+', sheet.cell(row=z, column=4).value)
                            if result[0] is None:
                                # обработать ошибку
                                break
                            for f in range(len(result)):
                                result[f] = int(result[f])
                            if amount < result[1]:
                                price = int(sheet.cell(row=z, column=5).value)
                                break
                            z = z + 1
                taxes_modifier = calculate_taxes(name)
                print(taxes_modifier, '- налоговый модификатор')
                print(amount, '- количество')
                print(price_modifier, '- модификатор цены')
                print(price, '- цена')
                final_price = amount * price_modifier * price
    price_per_unit = final_price / amount
    data_to_add = [name, amount, price_per_unit, final_price]
    data.append(data_to_add)
    print('final price', final_price)
    for i in data:
        print(i)
    a.withdraw()
    entry_calculate.delete(0, END)


# get list for top table
def get_list(value):
    my_list = []
    for j in range(1, rows + 1):
        if sheet.cell(row=j, column=value).value is None:
            j = j + 1
        else:
            my_list.append(sheet.cell(row=j, column=3).value)
    return my_list


root = Tk()
root.title("Советский калькулятор")

# some code
list_sampl = get_list_of_ids()
list_sampl = parse_list_of_ids(list_sampl)
dict_sample = get_dict_of_parents(list_sampl)
'''for i in dict_sample:
    print(i)'''

# main GUI params
box = Listbox(selectmode=EXTENDED, width=200, height=15)
box2 = Listbox(selectmode=EXTENDED, width=200, height=15)
default_list = get_list(3)
for i in default_list:
    box.insert(0, i)
scroll = Scrollbar(command=box.yview)
scroll2 = Scrollbar(command=box2.yview)
box.pack(side=TOP, expand=1, fill=X)
box2.pack(side=TOP, expand=1, fill=X)
box.config(yscrollcommand=scroll.set)
box2.config(yscrollcommand=scroll2.set)

# button frame
f = Frame()
f.pack(padx=180)
entry = Entry(f, width=50)
entry.pack(anchor=N)
Button(f, text="Найти", command=find_item) \
    .pack(fill=X)
Button(f, text="Рассчитать", command=calculate) \
    .pack(fill=X)
Button(f, text="Сохранить в файл", command=save_excel) \
    .pack(fill=X)

# event window
a = Toplevel()
a.title("Рассчет")
a.withdraw()
surname_label = Label(a, text="Количество:").pack()
entry_calculate = Entry(a, width=25)
entry_calculate.pack()
button3_calculate = Button(a, text="Сохранить", command=save_data) \
    .pack(fill=X)

root.mainloop()

workbook = xlsxwriter.Workbook('some_xlsx.xlsx')
worksheet = workbook.add_worksheet()
cell_format = workbook.add_format({'bold': True})
cell_format.set_font_size(20)
