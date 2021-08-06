from openpyxl import load_workbook
import re

my_table = load_workbook('infotecs_test.xlsx')
sheet_name = (str(my_table.active))[12:-2]
sheet = my_table[sheet_name]
rows = sheet.max_row
cols = sheet.max_column


def get_list_of_ids():
    my_list = []
    for i in range(1, rows + 1):
        if sheet.cell(row=i, column=2).value is None:
            i = i + 1
        else:
            my_list.append(sheet.cell(row=i, column=2).value)
    return my_list


def get_parent(some_list: list):
    child_list = []
    for i in range(len(some_list)):
        for j in range(1, rows + 1):
            if some_list[i] == sheet.cell(row=j, column=2).value:
                pattern = re.compile('\d{3}')
                result = pattern.search(str(sheet.cell(row=j, column=5).value))
                if result is not None:
                    child_list.append(some_list[i])
    return child_list


def get_list_of_sales():
    my_list = []
    for i in range(1, rows + 1):
        if sheet.cell(row=i, column=5).value is None:
            i = i + 1
        else:
            my_list.append(sheet.cell(row=i, column=5).value)
    return my_list


def parse_list_of_ids(list_sample: list):
    len_list = len(list_sample)
    for i in reversed(range(len_list)):
        result = re.search('\d{2}', list_sample[i])
        if result is None:
            list_sample.pop(i)
    len_list = len(list_sample)
    for i in reversed(range(len_list)):
        result1 = re.search(re.escape('*'), list_sample[i])
        if result1 is not None:
            list_sample.pop(i)
    len_list = len(list_sample)
    for i in reversed(range(len_list)):
        result1 = re.search(re.escape('960,00'), list_sample[i])
        if result1 is not None:
            list_sample.pop(i)
    list_sample.pop(0)
    return list_sample


def find_childs_by_id(product_id):
    list_sample = {}
    for i in range(1, rows + 1):
        pattern = re.compile(re.escape(product_id), flags=re.IGNORECASE)
        result = pattern.search(str(sheet.cell(row=i, column=2).value))
        print(result)
        try:
            if result[0] is not None:
                list_sample.update({str(sheet.cell(row=i, column=2).value): product_id})
        except:
            pass
    return list_sample


def get_dict_of_parents(list_of_id: list):
    list_of_id = get_parent(list_of_id)
    full_dict = {}
    for i in range(len(list_of_id)):
        update = find_childs_by_id(list_of_id[i])
        full_dict.update(update)
    return full_dict


def find_parent(child_name: str):
    list_sample = get_list_of_ids()
    list_sample = parse_list_of_ids(list_sample)
    dict_sample = get_dict_of_parents(list_sample)
    parent = dict_sample.get(child_name)
    return parent


x = find_parent('TLC-SW-04-G1')
print(x)

print(find_childs_by_id('HC-200-IDS2000-3.X'))
